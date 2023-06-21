from pathlib import Path
import requests
from lxml import etree
from openpyxl import Workbook, load_workbook
import json

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.63 Safari/537.36 Edg/102.0.1245.30'
}

head = ['车型', '官方指导价', '厂商', '级别', '能源类型', '上市时间',
        '电动机', '纯电续航里程(km)工信部', '纯电续航里程(km)CLTC',
        '充电时间(小时)', '快充电量(%)', '最大功率(kW)', '最大扭矩(N·m)',
        '变速箱', '长x宽x高(mm)', '车身结构', '最高车速(km/h)',
        '官方百公里加速时间(s)', '百公里耗电量(kWh/100km)',
        '电能当量燃料消耗量(L/100km)', '整车保修期限', '6万公里保养总成本预估']

head2 = ['车型', '官方指导价', '厂商', '级别', '能源类型', '上市时间',
        '发动机', '纯电续航里程(km)工信部', '纯电续航里程(km)NEDC',
        '纯电续航里程(km)WLTC', '综合续航里程(km)工信部', '充电时间(小时)',
         '快充电量(%)', '发动机最大功率(kW)', '电动机最大功率(kW)',
         '发动机最大扭矩(N·m)', '电动机最大扭矩(N·m)', '变速箱', '长x宽x高(mm)', '车身结构', '最高车速(km/h)',
        '官方百公里加速时间(s)', 'NEDC综合油耗(L/100km)', 'WLTC综合油耗(L/100km)',
         '百公里耗电量(kWh/100km)', '电能当量燃料消耗量(L/100km)',
         '最低荷电状态燃料消耗量(L/100km)', '整车保修期限', '6万公里保养总成本预估',
         '首任车主保修期限']


def main():
    url1 = "https://www.dongchedi.com/motor/pc/car/rank_data?aid=1839&app_name=auto_web_pc&city_" \
           "name=%E8%A5%BF%E5%AE%89&count=10&offset="

    url2 = "&month=&new_energy_type=1&rank_data_type=11&brand_id=&price=&manufacturer=&outter_detail_type=&nation=0"

    per_car_url1 = "https://www.dongchedi.com/auto/params-carIds-x-"

    all_cars = []
    for num in range(0, 160, 10):
        url = url1 + str(num) + url2
        res = requests.get(url=url, headers=headers)
        page_text = res.text
        page_text = json.loads(page_text)

        lists = page_text["data"]["list"]
        # 遍历每一个车的信息
        for per_car in lists:
            # 初始化每个车辆的信息
            car_params = {}
            for h in head:
                car_params[h] = None

            series_id = per_car["series_id"]
            series_name = per_car["series_name"]
            car_params['车型'] = series_name

            per_car_url = per_car_url1 + str(series_id)
            per_car_res = requests.get(url=per_car_url, headers=headers)
            per_car_page_text = per_car_res.text
            per_car_tree = etree.HTML(per_car_page_text)

            infos = per_car_tree.xpath('//div[@class="table_root__14vH_" and @name="config-body-0"]/div')
            price_key = infos[1].xpath('.//label[@class="cell_label__ZtXlw cell_has-wiki__18Gae"]/text()')[0]
            price = infos[1].xpath('.//div[@class="cell_official-price__1O2th"]/text()')[0]
            if price_key in car_params.keys():
                car_params[price_key] = price

            print("--"*4)
            for i in range(2, len(infos)):
                temps = infos[i].xpath('./div')
                print("len(temps) = ", len(temps))
                k = temps[0].xpath('.//label/text()')[0]
                v = temps[1].xpath('.//div[@class="cell_normal__37nRi"]/text()')
                v = v[0].replace("\n", " ") if v else None
                if k in car_params.keys():
                    car_params[k] = v
            for i, j in car_params.items():
                print(i, j)
            all_cars.append(car_params)
            manufacturer = infos[2].xpath('.//div[@class="cell_normal__37nRi"]/text()')[0]
            level = infos[3].xpath('.//div[@class="cell_normal__37nRi"]/text()')[0]
            energy_type = infos[4].xpath('.//div[@class="cell_normal__37nRi"]/text()')[0]
            # break
            print("--"*4)
        # break

    return all_cars


def proce_excel(cars):
    wb = Workbook()
    ws = wb.active

    ws.append(head)

    for car in cars:
        lines = []
        for h in head:
            lines.append(car[h])
        ws.append(lines)

    wb.save('test.xlsx')


def run():
    cars = main()
    proce_excel(cars)


def main2():
    url1 = "https://www.dongchedi.com/motor/pc/car/rank_data?aid=1839&app" \
           "_name=auto_web_pc&city_name=%E8%A5%BF%E5%AE%89&count=10&offset=" \

    url2 = "&month=&new_energy_type=2&rank_data_type=11&brand_id=&price=&manufacturer=&outter_detail_type=&nation=0"

    per_car_url1 = "https://www.dongchedi.com/auto/params-carIds-x-"

    all_cars = []
    for num in range(0, 50, 10):
        url = url1 + str(num) + url2
        res = requests.get(url=url, headers=headers)
        page_text = res.text
        page_text = json.loads(page_text)

        lists = page_text["data"]["list"]
        # 遍历每一个车的信息
        for per_car in lists:
            # 初始化每个车辆的信息
            car_params = {}
            for h in head2:
                car_params[h] = None

            series_id = per_car["series_id"]
            series_name = per_car["series_name"]
            car_params['车型'] = series_name

            per_car_url = per_car_url1 + str(series_id)
            per_car_res = requests.get(url=per_car_url, headers=headers)
            per_car_page_text = per_car_res.text
            per_car_tree = etree.HTML(per_car_page_text)

            infos = per_car_tree.xpath('//div[@class="table_root__14vH_" and @name="config-body-0"]/div')
            price_key = infos[1].xpath('.//label[@class="cell_label__ZtXlw cell_has-wiki__18Gae"]/text()')[0]
            price = infos[1].xpath('.//div[@class="cell_official-price__1O2th"]/text()')[0]
            if price_key in car_params.keys():
                car_params[price_key] = price

            print("--"*4)
            for i in range(2, len(infos)):
                temps = infos[i].xpath('./div')
                k = temps[0].xpath('.//label/text()')[0]
                v = temps[1].xpath('.//div[@class="cell_normal__37nRi"]/text()')
                v = v[0].replace("\n", " ") if v else None
                if k in car_params.keys():
                    car_params[k] = v
            for i, j in car_params.items():
                print(i, j)
            all_cars.append(car_params)
            manufacturer = infos[2].xpath('.//div[@class="cell_normal__37nRi"]/text()')[0]
            level = infos[3].xpath('.//div[@class="cell_normal__37nRi"]/text()')[0]
            energy_type = infos[4].xpath('.//div[@class="cell_normal__37nRi"]/text()')[0]
            print("--"*4)
            # break
        # break

    return all_cars


def proce_excel2(cars):
    wb = Workbook()
    ws = wb.active

    ws.append(head2)

    for car in cars:
        lines = []
        for h in head2:
            lines.append(car[h])
        ws.append(lines)

    wb.save('test2.xlsx')


def run2():
    cars = main2()
    proce_excel2(cars)


if __name__ == "__main__":
    # main()
    run()
    run2()
    # countries()
