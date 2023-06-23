from pathlib import Path
import tushare as ts
import pandas as pd
import time
import numpy as np
from openpyxl import Workbook, load_workbook

# 四个行业
industries = ['电子', '通信', '传媒', '计算机']

# 保存四个行业公司股票代码的excel路径
excel_path = r'D:\workspace\code\study\py_other\四行业数据.xlsx'


class Run:
    def __init__(self, token, start, end):
        self.token = token
        self.start = start
        self.end = end

    def get_datas(self):
        infos = self.preprocess()
        pro = ts.pro_api(token=self.token)

        # 遍历每个行业
        for industry_name, codes in infos.items():
            industry_closes = []
            industry_highs = []
            industry_lows = []
            # 遍历每个公司的股票代码
            for code in codes:
                df = ts.get_hist_data(code=code, start=self.start, end=self.end)

                df = df.to_dict()
                # 收盘价
                closes = df['close']
                industry_closes.append(closes)

                # 最高价
                highs = df['high']
                industry_highs.append(highs)

                # 最低价
                lows = df['low']
                industry_lows.append(lows)

                assert len(closes) == len(highs) == len(lows), "three numbers should be same"
                # break
            # for c, h, l in zip(industry_closes, industry_highs, industry_lows):
            #     print(len(c), len(h), len(l))
            industry_closes, industry_highs, industry_lows = \
                self.alignment(industry_closes, industry_highs, industry_lows)
            # print('***', len(industry_closes), len(industry_highs), len(industry_lows))
            # print(industry_lows[-1])

            new_industry_closes, new_industry_highs, new_industry_lows = [], [], []
            for c, h, l in zip(industry_closes, industry_highs, industry_lows):
                temp = []
                temp2 = []
                temp3 = []
                for k in c.keys():
                    temp.append(c[k])
                    temp2.append(h[k])
                    temp3.append(l[k])

                new_industry_closes.append(temp)
                new_industry_highs.append(temp2)
                new_industry_lows.append(temp3)

            # [公司数, 日股票信息], 如 [100, 607] 就表示有100家公司，每家公司的股票信息是607个
            print(np.array(new_industry_closes).shape,
                  np.array(new_industry_highs).shape,
                  np.array(new_industry_lows).shape)
            self.cal_nh_nl(new_industry_closes, new_industry_highs, new_industry_lows)
            break

    @staticmethod
    def alignment(industry_closes, industry_highs, industry_lows):
        """
        同一行业下的公司，因为上市时间有早晚，可能会导致同一时间段内的股票信息数量不一致
        所以该函数就是做数据对齐操作
        Returns:

        """
        new_industry_closes = []
        new_industry_highs = []
        new_industry_lows = []
        arr = [len(c) for c in industry_closes]
        max_value = int(np.max(arr))
        max_value_idx = int(np.argmax(arr))
        for c, h, l in zip(industry_closes, industry_highs, industry_lows):
            if len(c) == max_value:
                new_industry_closes.append(c)
                new_industry_highs.append(h)
                new_industry_lows.append(l)
            # 给缺失日期补零
            else:
                temp = {}
                temp2 = {}
                temp3 = {}
                for k in industry_closes[max_value_idx].keys():
                    if k in c.keys():
                        temp[k] = c[k]
                        temp2[k] = h[k]
                        temp3[k] = l[k]
                    else:
                        temp[k] = 0
                        temp2[k] = 0
                        temp3[k] = 0
                new_industry_closes.append(temp)
                new_industry_highs.append(temp2)
                new_industry_lows.append(temp3)

        return new_industry_closes, new_industry_highs, new_industry_lows

    @staticmethod
    def preprocess():
        """
        预处理excel文件
        Returns:
            dict {行业名称: [股票代码, ...]}
        """
        infos = {}
        wb = load_workbook(filename=excel_path)
        for i in industries:
            infos[i] = []
            ws = wb[i]
            for row_idx in range(2, ws.max_row + 1):
                loc = 'A' + str(row_idx)
                v = str(ws[loc].value).rjust(6, '0')
                infos[i].append(v)

        # for k, v in infos.items():
        #     print(k, len(v))

        return infos

    @staticmethod
    def cal_nh_nl(industry_closes, industry_highs, industry_lows):
        """
        计算行业净新高占比
        Returns:

        """
        for c in industry_closes:
            print(c, len(c))


def main():
    start = time.time()
    t = '818670fa68bc204c217143cdb75efeae1986031841ff8ca2c6a855bd'
    s = '20100501'
    e = '20230430'
    run = Run(token=t, start=s, end=e)
    run.get_datas()
    # arr = [10, 4, 10, 5, 10]
    # print(np.max(arr), type(np.argmax(arr)))
    cost_time = time.time() - start
    print(f"耗时 {cost_time:.2f} s")


if __name__ == "__main__":
    main()
    # preprocess()



