from openpyxl import Workbook, load_workbook
from openpyxl.utils import FORMULAE
import datetime


def main():
    wb = Workbook()
    ws = wb.active

    ws.append(['价格1', '价格2', '总和', '均值'])

    ws.append([22, 63])
    ws.append([11, 88])
    ws.append([15, 68])

    ws["c2"] = "=SUM(A2, B2)"   # 求和
    ws["d2"] = "=AVERAGE(A2: B2)"  # 求平均值

    wb.save('test.xlsx')


if __name__ == "__main__":
    main()


